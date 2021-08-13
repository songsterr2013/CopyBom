import os
import sys
import shutil
from datetime import date
from openpyxl import load_workbook
import util

util.running_prerequisite()
logger = util.get_logger(__file__)


class CopyBom:

    root_path = ''
    target_path = ''

    def __init__(self, bom_path, bom_list_file):

        self.bom_path = bom_path  # 集中放BOM的路徑
        self.file_path = bom_list_file  # 待測試的檔案

        self.wb = load_workbook(bom_list_file, read_only=False)
        self.sh = self.wb[self.wb.sheetnames[0]]

    # 把檔案裡面的BOM yield出來
    def parse_bom(self):
        table = self.sh
        for row in table.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0] is not None:
                yield row[0][0], row[0]

    # 複製檔案
    def action(self, prefix, file_name):
        path = os.path.join(self.bom_path, prefix, file_name)
        file_format = self.check_format(path)
        if file_format is not None:
            completed_path = os.path.join(self.bom_path, prefix, file_name + file_format)
            final_location = os.path.join(self.target_path, file_name + file_format)
            shutil.copyfile(completed_path, final_location)
        else:
            logger.info("沒有該檔案:{}".format(file_name))

    # 創folder
    def make_dir(self):
        self.get_previous_path()

        today = str(date.today())
        try:
            self.target_path = os.path.join(self.root_path, '{}'.format(today + '_testing_bom'))
            os.mkdir(self.target_path)
        except:
            pass

    def get_previous_path(self):

        if getattr(sys, 'frozen', False):  # 如果為exe
            self.root_path = os.path.dirname(sys.executable)
        elif __file__:
            self.root_path = os.path.dirname(__file__)

    @staticmethod  # return一個檔案格式
    def check_format(path):
        if os.path.isfile(path + '.xls'):
            return '.xls'
        elif os.path.isfile(path + '.xlsx'):
            return '.xlsx'
        else:
            return None

    def save(self):
        self.wb.save(self.file_path)
